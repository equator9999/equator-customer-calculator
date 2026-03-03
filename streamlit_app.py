import os
import sys
import json
import tempfile
import subprocess
from pathlib import Path
import io
import re

import streamlit as st
import pandas as pd

import invoice_audit  # NEW/UPDATED FILE


# --- Ensure required packages exist in THIS environment (Streamlit Cloud can be flaky) ---
def ensure(pkg: str) -> None:
    try:
        __import__(pkg)
    except Exception:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])


ensure("openpyxl")
ensure("pandas")
ensure("numpy")
ensure("requests")
ensure("pycountry")
ensure("pdfplumber")  # needed for DHL + FEDEX PDF parsing

import openpyxl
import pycountry
import run_pricing


APP_DIR = Path(__file__).parent
RATES_FILE = APP_DIR / "RETOOL ALL COST UPLOAD 2026 WITH FUEL TYPE.xlsx"
TEMPLATE_FILE = APP_DIR / "Shipment Template.xlsx"

TRANSIT_BASE = APP_DIR / "TRANSIT_BASE.xlsx"
TRANSIT_OVERRIDES = APP_DIR / "TRANSIT_CITY_OVERRIDES.xlsx"
CITYCLASS_MASTER = APP_DIR / "CITYCLASS_MASTER.xlsx"
HS_DB = APP_DIR / "hs_codes.csv"

LOGO_FILE = APP_DIR / "logo.png"

st.set_page_config(page_title="Pricing Calculator", layout="wide")


# ---------------------------
# Invoice dest cleanup helpers
# ---------------------------
UK_POSTCODE_RE = re.compile(r"\b([A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2})\b", re.I)


def extract_dest_from_combo(raw: str, fallback_country: str = "") -> tuple[str, str]:
    """
    Handles strings like:
      'En3 7Qa Enfield Nl-5624 Eb Eindhoven' -> ('NL', 'Eindhoven')
      'En8 0Ag Hertfordshire Cy-2800 Kakopetria' -> ('CY', 'Kakopetria')
      '... Gb-Bt8 8Nh Carryduff' -> ('GB', 'Carryduff')
    Looks for LAST occurrence of 'CC-XXXX' and keeps what comes after it as city-ish.
    """
    if raw is None:
        return (str(fallback_country or "").upper(), "")
    s = str(raw).strip()
    if not s:
        return (str(fallback_country or "").upper(), "")

    s = re.sub(r"\s+", " ", s).strip()

    matches = list(re.finditer(r"\b([A-Z]{2})-([A-Z0-9]{2,})\b", s, flags=re.I))
    if matches:
        m = matches[-1]
        cc = m.group(1).upper()
        tail = s[m.end():].strip()

        city_guess = tail
        city_guess = UK_POSTCODE_RE.sub("", city_guess).strip()
        city_guess = re.sub(r"\b\d{4}\s*[A-Z]{2}\b", "", city_guess, flags=re.I).strip()  # NL 5624 EB
        city_guess = re.sub(r"\b\d{3,6}\b", "", city_guess).strip()                        # CY 2800 etc
        city_guess = re.sub(r"\s+", " ", city_guess).strip()

        return cc, city_guess

    cleaned = UK_POSTCODE_RE.sub("", s).strip()
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return (str(fallback_country or "").upper(), cleaned)


def iso2_to_country_name(x: str) -> str:
    """
    Convert ISO2 (GB/NL/CY) -> country name likely used in ZONES sheet.
    If already a name, returns original.
    """
    if x is None:
        return ""
    s = str(x).strip()
    if not s:
        return ""
    if len(s) == 2 and s.isalpha():
        a2 = s.upper()
        # common special cases (match typical workbook naming)
        if a2 == "GB":
            return "UNITED KINGDOM"
        if a2 == "US":
            return "UNITED STATES"
        try:
            c = pycountry.countries.get(alpha_2=a2)
            if c and getattr(c, "name", None):
                return str(c.name).upper()
        except Exception:
            pass
        return a2
    return s


# ---------- Header ----------
header_cols = st.columns([1.2, 6, 1])
with header_cols[0]:
    if LOGO_FILE.exists():
        st.image(str(LOGO_FILE), use_container_width=True)
with header_cols[1]:
    st.markdown("## Pricing Calculator (Staff)")
    st.caption("Shipments pricing + Invoice audit (Model A)")
with header_cols[2]:
    st.write("")


# Password gate (set APP_PASSWORD in Streamlit Secrets)
PASSWORD = os.environ.get("APP_PASSWORD", "")
if PASSWORD:
    entered = st.text_input("Password", type="password", key="pw")
    if entered != PASSWORD:
        st.stop()


# ---------------------------
# Caching helpers
# ---------------------------
@st.cache_data(show_spinner=False)
def read_fuel_table_cached(rates_path_str: str) -> pd.DataFrame | None:
    rates_path = Path(rates_path_str)
    try:
        wb = openpyxl.load_workbook(rates_path, data_only=True)
        if "FUEL" not in wb.sheetnames:
            return None
        ws = wb["FUEL"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).strip() if h is not None else "" for h in headers]
        if not any(headers):
            return None
        rows = []
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            rows.append(row)
        return pd.DataFrame(rows, columns=headers)
    except Exception:
        return None


def normalise_pct_input(x: float) -> float:
    if x is None:
        return 0.0
    v = float(x)
    if v > 1.0:
        v = v / 100.0
    return v


def ensure_to_city_in_template_bytes(template_bytes: bytes) -> bytes:
    bio = io.BytesIO(template_bytes)
    wb = openpyxl.load_workbook(bio)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers_str = [str(h).strip() if h is not None else "" for h in headers]

    if "To City" not in headers_str:
        try:
            idx = headers_str.index("To Country") + 1
        except ValueError:
            idx = len(headers_str)
        ws.insert_cols(idx + 1, 1)
        ws.cell(1, idx + 1).value = "To City"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def style_hs_risk(df: pd.DataFrame):
    if df is None or df.empty:
        return df.style

    def _risk_style(val):
        v = str(val).strip().upper()
        if v == "LOW":
            return "background-color: #d7f7d7; color: #0b4f0b; font-weight: 700;"
        if v in ("MED", "MEDIUM"):
            return "background-color: #fff2cc; color: #7a4b00; font-weight: 700;"
        if v == "HIGH":
            return "background-color: #ffd6d6; color: #7a0000; font-weight: 700;"
        return ""

    if "HS_Risk" in df.columns:
        return df.style.applymap(_risk_style, subset=["HS_Risk"])
    return df.style


# ---------------------------
# Required file checks
# ---------------------------
missing = []
for f in [RATES_FILE, TRANSIT_BASE, TRANSIT_OVERRIDES, CITYCLASS_MASTER]:
    if not f.exists():
        missing.append(f.name)

if missing:
    st.error("Missing required files in repo root: " + ", ".join(missing))
    st.stop()


# ---------------------------
# 1) Template download
# ---------------------------
st.subheader("1) Shipment template")

colA, colB = st.columns([2, 3])
with colA:
    if TEMPLATE_FILE.exists():
        templ_bytes = ensure_to_city_in_template_bytes(TEMPLATE_FILE.read_bytes())
        st.download_button(
            label="Download Shipment Template (includes To City)",
            data=templ_bytes,
            file_name="Shipment Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.warning('Template file not found. Ensure "Shipment Template.xlsx" exists in the repo root.')

with colB:
    st.info("Tip: Invoice audit does NOT require the shipment template.", icon="ℹ️")


# ---------------------------
# 2) Mode
# ---------------------------
st.subheader("2) What are you doing today?")

mode = st.radio(
    "Mode",
    ["Price Shipments", "Invoice Audit (Model A)"],
    horizontal=True,
    index=0,
)

carrier_selector = st.selectbox(
    "Carrier (for Invoice Audit)",
    ["DHL", "UPS", "FEDEX"],
    index=0,
    help="Supported today: DHL PDF, FEDEX PDF. UPS coming next.",
)

st.divider()


# ---------------------------
# Common inputs form
# ---------------------------
with st.form("inputs_form", clear_on_submit=False):
    st.subheader("3) Filters")

    f1, f2, f3 = st.columns([2, 2, 2])
    with f1:
        carriers_selected = st.multiselect(
            "Carriers (pricing engine filter)",
            options=["UPS", "DHL", "FEDEX"],
            default=["UPS", "DHL", "FEDEX"],
        )
    with f2:
        types_selected = st.multiselect(
            "Service types",
            options=["EXPRESS", "ECONOMY", "DOMESTIC", "FREIGHT"],
            default=["EXPRESS", "ECONOMY", "DOMESTIC", "FREIGHT"],
        )
    with f3:
        st.caption("Invoice Audit automatically restricts engine to the selected carrier.")

    st.subheader("4) Customs (optional)")
    include_customs = st.checkbox("Include duties & taxes (Easyship)", value=False)
    only_cheapest_customs = False
    cheapest_n = 5
    if include_customs:
        only_cheapest_customs = st.checkbox("Only call duties/taxes for cheapest options", value=True)
        if only_cheapest_customs:
            cheapest_n = st.number_input("Cheapest N per shipment (customs calls)", 1, 25, 5, 1)

    st.subheader("5) Defaults (for missing fields)")
    d1, d2, d3, d4 = st.columns([1.2, 1.2, 1.2, 2.2])
    with d1:
        default_origin = st.text_input("Default Origin (ISO2)", value="GB")
    with d2:
        default_currency = st.selectbox("Default Currency", ["GBP", "EUR", "USD"], index=0)
    with d3:
        default_incoterm = st.selectbox("Default Incoterm", ["DAP", "DDP"], index=0)
    with d4:
        default_hs = st.text_input("Default HS Code", value="49111090")
    default_item_value = st.number_input("Default item value (GBP)", min_value=0.0, value=10.0, step=1.0)

    st.subheader("6) Fuel overrides (optional)")
    fuel_df = read_fuel_table_cached(str(RATES_FILE))
    with st.expander("Show current fuel table", expanded=False):
        if fuel_df is None:
            st.info("No readable FUEL sheet found.")
        else:
            st.dataframe(fuel_df, use_container_width=True, height=220)

    override_fuel = st.checkbox("Override fuel % (this run only)", value=False)
    fuel_overrides = {}
    if override_fuel:
        carriers = ["UPS", "DHL", "FEDEX"]
        types = ["EXPRESS", "ECONOMY", "DOMESTIC", "FREIGHT"]

        prefill = {}
        if fuel_df is not None and {"Carrier", "Type", "FuelPct"}.issubset(set(fuel_df.columns)):
            for _, r in fuel_df.iterrows():
                c = str(r.get("Carrier", "")).strip().upper()
                t = str(r.get("Type", "")).strip().upper()
                p = r.get("FuelPct", None)
                try:
                    if p is not None and str(p).strip() != "":
                        p = float(p)
                        if p > 1:
                            p = p / 100.0
                        prefill[(c, t)] = p
                except Exception:
                    pass

        g = st.columns(5)
        g[0].markdown("**Carrier**")
        for i, t in enumerate(types, start=1):
            g[i].markdown(f"**{t[:4]}**")

        for c in carriers:
            row_cols = st.columns(5)
            row_cols[0].markdown(f"**{c}**")
            for i, t in enumerate(types, start=1):
                default_val = prefill.get((c, t), None)
                default_display = (default_val * 100.0) if isinstance(default_val, (int, float)) else 0.0
                val = row_cols[i].number_input(
                    label=f"{c}_{t}",
                    min_value=0.0,
                    max_value=200.0,
                    value=float(default_display),
                    step=0.1,
                    key=f"fuel_{c}_{t}",
                    label_visibility="collapsed",
                )
                if val and float(val) > 0:
                    fuel_overrides[f"{c}|{t}"] = normalise_pct_input(val)

    st.subheader("7) Upload + run")

    divisor = st.number_input("Volumetric divisor (cm)", 1000, 20000, 5000, 100)
    max_parcels = st.number_input("Max parcels per row", 1, 5, 5, 1)

    if mode == "Price Shipments":
        upload = st.file_uploader("Shipment file (xlsx/xlsm)", type=["xlsx", "xlsm"])
        run_btn = st.form_submit_button("Calculate Prices", type="primary")
        invoice_upload = None
    else:
        invoice_upload = st.file_uploader("Carrier invoice (PDF)", type=["pdf"])
        run_btn = st.form_submit_button("Run Invoice Audit", type="primary")
        upload = None


# ---------------------------
# Engine runner
# ---------------------------
def run_engine_on_shipments(df: pd.DataFrame, tmpdir: Path):
    ship_path = tmpdir / "shipments_normalised.xlsx"
    out_path = tmpdir / "pricing_output.xlsx"

    for col in ["Origin Country", "Currency", "Incoterm", "To City"]:
        if col not in df.columns:
            df[col] = ""

    def _fill_default(col, default_val):
        df[col] = df[col].fillna("").astype(str)
        df.loc[df[col].str.strip().isin(["", "nan", "None"]), col] = str(default_val)

    _fill_default("Origin Country", default_origin)
    _fill_default("Currency", default_currency)
    _fill_default("Incoterm", default_incoterm)

    for i in range(1, int(max_parcels) + 1):
        hs_col = f"P{i}_HSCode"
        val_col = f"P{i}_Value"
        if hs_col not in df.columns:
            df[hs_col] = ""
        if val_col not in df.columns:
            df[val_col] = ""

        df[hs_col] = df[hs_col].fillna("").astype(str)
        df[val_col] = df[val_col].fillna("").astype(str)

        df.loc[df[hs_col].str.strip().isin(["", "nan", "None"]), hs_col] = str(default_hs).strip()
        df.loc[df[val_col].str.strip().isin(["", "nan", "None"]), val_col] = str(float(default_item_value))

    df.to_excel(ship_path, index=False)

    fuel_overrides_json = json.dumps(fuel_overrides) if fuel_overrides else ""

    old_argv = sys.argv[:]
    try:
        sys.argv = [
            "run_pricing.py",
            "--rates", str(RATES_FILE),
            "--shipments", str(ship_path),
            "--out", str(out_path),
            "--divisor", str(int(divisor)),
            "--max_parcels", str(int(max_parcels)),
            "--carriers", ",".join(carriers_selected) if carriers_selected else "",
            "--types", ",".join(types_selected) if types_selected else "",
            "--include_customs", "1" if include_customs else "0",
            "--customs_only_cheapest", "1" if (include_customs and only_cheapest_customs) else "0",
            "--customs_top_n", str(int(cheapest_n)),
            "--fuel_overrides_json", fuel_overrides_json,
        ]
        run_pricing.main()
    finally:
        sys.argv = old_argv

    results = pd.read_excel(out_path, sheet_name="Prices_All_Services")
    return out_path, results


# ---------------------------
# Run button handler
# ---------------------------
if run_btn:
    with st.spinner("Running…"):
        with tempfile.TemporaryDirectory() as td:
            td = Path(td)

            if mode == "Price Shipments":
                if upload is None:
                    st.error("Please upload a shipment file.")
                    st.stop()

                df = pd.read_excel(upload, dtype=str)
                out_path, results = run_engine_on_shipments(df, td)

                num_cols = [
                    "ChargeableKg","BaseFreight","OversizeAmount","NonConvAmount","OverGirthAmount",
                    "SubtotalNonFuel","SubtotalBeforeFuel","FuelAmount","FinalTotal",
                    "GoodsValueTotal","DutyTotal","TaxTotal","LandedCostTotal",
                ]
                for c in num_cols:
                    if c in results.columns:
                        results[c] = pd.to_numeric(results[c], errors="coerce").round(2)

                show_cols = [
                    "Shipment ID","Country","ToCity",
                    "CarrierKey","Service","Type",
                    "FinalTotal","IsLowestRatePerShipment",
                    "TransitDays_Min","TransitDays_Max","TransitSource","CityClass",
                    "HSCode_Used","HS_Risk",
                    "FuelPctSource",
                ]
                show_cols = [c for c in show_cols if c in results.columns]
                preview = results[show_cols].copy()

                st.success("Done! Preview below and download.")
                st.dataframe(style_hs_risk(preview), use_container_width=True, height=520)

                st.download_button(
                    "Download pricing_output.xlsx",
                    data=out_path.read_bytes(),
                    file_name="pricing_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            else:
                if invoice_upload is None:
                    st.error("Please upload an invoice PDF.")
                    st.stop()

                inv_bytes = invoice_upload.read()
                carrier = carrier_selector

                try:
                    parsed = invoice_audit.parse_invoice(carrier, invoice_upload.name, inv_bytes)
                except Exception as e:
                    st.error(str(e))
                    st.stop()

                if parsed.warnings:
                    for w in parsed.warnings:
                        st.warning(w)

                inv_ship = parsed.shipments.copy()
                inv_lines = parsed.surcharge_lines.copy()

                if inv_ship.empty:
                    st.error("Parsed invoice but found no shipments. Parser may need tweak for this layout.")
                    st.stop()

                # --- FIX combined DestCity into DestCountry + DestCity ---
                if "DestCountry" not in inv_ship.columns:
                    inv_ship["DestCountry"] = ""
                if "DestCity" not in inv_ship.columns:
                    inv_ship["DestCity"] = ""

                inv_ship["DestCountry"] = inv_ship["DestCountry"].astype(str).str.upper()

                def _fix_row(r):
                    cc, city = extract_dest_from_combo(r.get("DestCity", ""), r.get("DestCountry", ""))
                    return pd.Series({"DestCountry": cc, "DestCity": city})

                fixed = inv_ship.apply(_fix_row, axis=1)
                inv_ship["DestCountry"] = fixed["DestCountry"]
                inv_ship["DestCity"] = fixed["DestCity"]

                # --- IMPORTANT: convert ISO2 -> country name for engine lookup ---
                inv_ship["DestCountry_ForEngine"] = inv_ship["DestCountry"].apply(iso2_to_country_name)

                # Build shipment df for engine
                ship_df = pd.DataFrame()
                ship_df["Shipment ID"] = inv_ship["ShipmentKey"].astype(str)
                ship_df["To Country"] = inv_ship["DestCountry_ForEngine"].astype(str)
                ship_df["To City"] = inv_ship["DestCity"].astype(str)
                ship_df["Origin Country"] = default_origin
                ship_df["Currency"] = default_currency
                ship_df["Incoterm"] = default_incoterm

                # Use billed weight as P1_DeadKg
                ship_df["P1_DeadKg"] = inv_ship["BilledWeightKg"].fillna(1).astype(str)

                for i in range(1, int(max_parcels) + 1):
                    if f"P{i}_DeadKg" not in ship_df.columns:
                        ship_df[f"P{i}_DeadKg"] = ""
                    if f"P{i}_L" not in ship_df.columns:
                        ship_df[f"P{i}_L"] = ""
                    if f"P{i}_W" not in ship_df.columns:
                        ship_df[f"P{i}_W"] = ""
                    if f"P{i}_H" not in ship_df.columns:
                        ship_df[f"P{i}_H"] = ""
                    if f"P{i}_HSCode" not in ship_df.columns:
                        ship_df[f"P{i}_HSCode"] = default_hs
                    if f"P{i}_Value" not in ship_df.columns:
                        ship_df[f"P{i}_Value"] = str(float(default_item_value))

                # Restrict engine to this carrier for fairness
                carriers_selected = [carrier_selector]

                out_path, model_results = run_engine_on_shipments(ship_df, td)

                # keep only carrier rows
                if "CarrierKey" in model_results.columns:
                    model_results = model_results[model_results["CarrierKey"].astype(str).str.upper() == carrier_selector.upper()].copy()

                # Harden schema
                for col in ["Shipment ID", "Type", "Service", "FinalTotal", "FuelAmount", "BaseFreight", "OversizeAmount", "NonConvAmount", "OverGirthAmount"]:
                    if col not in model_results.columns:
                        model_results[col] = "" if col in ("Shipment ID", "Type", "Service") else pd.NA

                model_results["Shipment ID"] = model_results["Shipment ID"].astype(str)
                model_results["Type"] = model_results["Type"].astype(str).str.upper()
                model_results["FinalTotal"] = pd.to_numeric(model_results["FinalTotal"], errors="coerce")

                inv_ship["ShipmentKey"] = inv_ship["ShipmentKey"].astype(str)
                inv_ship["Type"] = inv_ship["Type"].astype(str).str.upper()

                join_type = inv_ship[["ShipmentKey", "Type"]].rename(columns={"ShipmentKey": "Shipment ID", "Type": "Type_Invoice"})
                model_typed = model_results.merge(join_type, on="Shipment ID", how="left")

                if "Type" not in model_typed.columns:
                    model_typed["Type"] = ""
                if "Type_Invoice" not in model_typed.columns:
                    model_typed["Type_Invoice"] = ""

                model_typed["Type"] = model_typed["Type"].astype(str).str.upper()
                model_typed["Type_Invoice"] = model_typed["Type_Invoice"].astype(str).str.upper()
                model_match = model_typed[model_typed["Type"] == model_typed["Type_Invoice"]].copy()

                if not model_match.empty:
                    idx = model_match.groupby("Shipment ID")["FinalTotal"].idxmin()
                    best_model = model_match.loc[idx].copy()
                else:
                    best_model = pd.DataFrame(columns=[
                        "Shipment ID","Service","Type","FinalTotal","FuelAmount","BaseFreight","OversizeAmount","NonConvAmount","OverGirthAmount"
                    ])

                rec = inv_ship.rename(columns={"ShipmentKey": "Shipment ID"}).merge(
                    best_model[["Shipment ID","Service","Type","FinalTotal","FuelAmount","BaseFreight","OversizeAmount","NonConvAmount","OverGirthAmount"]],
                    on="Shipment ID",
                    how="left",
                    suffixes=("_Invoice", "_Model"),
                )

                rec = rec.rename(columns={
                    "FinalTotal": "Model_Total",
                    "FuelAmount": "Model_Fuel",
                    "BaseFreight": "Model_Base",
                    "OversizeAmount": "Model_Oversize",
                    "NonConvAmount": "Model_NonConv",
                    "OverGirthAmount": "Model_OverGirth",
                    "Service": "Model_Service",
                })

                rec["Invoice_TotalExVAT"] = pd.to_numeric(rec["Invoice_TotalExVAT"], errors="coerce")
                rec["Model_Total"] = pd.to_numeric(rec["Model_Total"], errors="coerce")

                rec["Delta"] = (rec["Invoice_TotalExVAT"] - rec["Model_Total"]).round(2)
                rec["DeltaPct"] = ((rec["Delta"] / rec["Model_Total"]) * 100.0).round(2)

                def driver(row):
                    if pd.isna(row["Model_Total"]):
                        return "NO_MODEL_MATCH"
                    if abs(float(row["Delta"])) < 0.01:
                        return "MATCH"
                    if float(row.get("Invoice_Fuel", 0) or 0) > 0 and (pd.isna(row.get("Model_Fuel")) or float(row.get("Model_Fuel") or 0) == 0):
                        return "FUEL_MISSING_IN_MODEL"
                    if float(row.get("Invoice_Surcharges", 0) or 0) > 0:
                        return "SURCHARGE_MISMATCH"
                    return "RATE_MISMATCH"

                rec["DeltaDriver"] = rec.apply(driver, axis=1)

                rec["AbsDelta"] = rec["Delta"].abs()
                rec = rec.sort_values(["AbsDelta"], ascending=False).drop(columns=["AbsDelta"])

                show_cols = [
                    "Shipment ID", "DestCountry", "DestCity", "Type",
                    "Invoice_Base", "Invoice_Fuel", "Invoice_Surcharges", "Invoice_TotalExVAT", "InvoiceTotalSource",
                    "Model_Service", "Model_Total",
                    "Delta", "DeltaPct", "DeltaDriver",
                ]
                show_cols = [c for c in show_cols if c in rec.columns]
                preview = rec[show_cols].copy()

                st.success("Invoice audit complete. Preview below and download.")
                st.dataframe(preview, use_container_width=True, height=560)

                audit_xlsx = td / "invoice_audit_output.xlsx"
                with pd.ExcelWriter(audit_xlsx, engine="openpyxl") as writer:
                    inv_ship.to_excel(writer, index=False, sheet_name="Invoice_Shipments")
                    inv_lines.to_excel(writer, index=False, sheet_name="Invoice_ChargeLines")
                    model_results.to_excel(writer, index=False, sheet_name="Model_AllRows")
                    rec.to_excel(writer, index=False, sheet_name="Reconciliation")

                st.download_button(
                    "Download invoice_audit_output.xlsx",
                    data=audit_xlsx.read_bytes(),
                    file_name="invoice_audit_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
